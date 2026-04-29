import streamlit as st
import pandas as pd
import re
import json
import os
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, timedelta
from streamlit_gsheets import GSheetsConnection
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook, load_workbook
from io import BytesIO

# --- 1. CONFIGURAÇÕES TÉCNICAS ---
ARQUIVO_TAGS = "tags_salvas.json"
ARQUIVO_CIDADES = "cidades.xlsx"

# --- CONEXÃO INICIAL ---
conn = st.connection("gsheets", type=GSheetsConnection)

def carregar_usuarios():
    try:
        # TTL de 1s para garantir que novos usuários apareçam logo
        df = conn.read(worksheet="usuários", ttl="1s")
        return df.dropna(how='all')
    except:
        return pd.DataFrame(columns=["usuario", "senha", "nivel"])

def carregar_tags():
    padrao = {"tags": {}, "last_selection": {}}
    if os.path.exists("tags_salvas.json"):
        try:
            with open("tags_salvas.json", "r", encoding="utf-8") as f:
                return json.load(f)
        except: 
            return padrao
    return padrao


# --- ESTADOS DE SESSÃO ---
if "logado" not in st.session_state:
    st.session_state.logado = False
    st.session_state.usuario_ativo = None
    st.session_state.nivel_ativo = None

# --- DEFINIÇÃO DO CAMINHO DA LOGO ---
diretorio_atual = os.path.dirname(os.path.abspath(__file__))
caminho_logo = os.path.join(diretorio_atual, "logo.png")

# --- CONFIGURAÇÕES DA PÁGINA ---
st.set_page_config(
    page_title="SISTEMA ADM | PROFISSIONALIZA", 
    layout="wide", 
    initial_sidebar_state="collapsed",
    page_icon=caminho_logo if os.path.exists(caminho_logo) else None
)

# --- CSS HUD NEON ---
st.markdown("""
    <style>
    .stApp { background-color: #0b0e1e; color: #e0e0e0; }
    .stTabs [data-baseweb="tab-list"] { 
        background-color: #121629; border-bottom: 1px solid #1f295a;
        position: fixed; top: 0; left: 0 !important; width: 100vw !important;
        z-index: 999; justify-content: center; height: 35px !important;
    }
    .stTabs [data-baseweb="tab"] { color: #64748b !important; font-size: 11px !important; padding: 0 30px !important; }
    .stTabs [aria-selected="true"] { color: #00f2ff !important; border-bottom: 2px solid #00f2ff !important; background-color: rgba(0, 242, 255, 0.05) !important; }
    
    .main .block-container { padding-top: 40px !important; max-width: 100% !important; margin: 0 auto !important; }
    
    label { color: #00f2ff !important; font-weight: bold !important; font-size: 17px !important; display: flex; align-items: center; justify-content: flex-end; }
    
    /* CADASTRO EM MAIÚSCULO */
    .stTabs div[data-testid="stTextInput"] input { text-transform: uppercase !important; }
    
    /* LOGIN EM MINÚSCULO E ESTILO */
    .login-box div[data-testid="stTextInput"] input { text-transform: none !important; background-color: white !important; color: black !important; }
    
    .login-box { 
        background: rgba(18, 22, 41, 0.9); padding: 40px; border-radius: 15px; 
        border: 2px solid #1f295a; box-shadow: 0 0 30px rgba(0, 242, 255, 0.1);
        margin-top: 100px; text-align: center;
    }
    
    .login-box button { background-color: #bc13fe !important; color: white !important; font-weight: bold !important; }
    .stTextInput input { background-color: white !important; color: black !important; font-size: 12px !important; height: 18px !important; border-radius: 5px !important; }
    .stCheckbox label p { color: #2ecc71 !important; font-weight: bold !important; font-size: 11px !important; }

    .custom-table-wrapper { width: 100%; max-height: 600px; overflow: auto; background-color: #121629; border: 2px solid #1f295a; border-radius: 10px; margin-top: 15px; }
    .custom-table { width: 100%; border-collapse: collapse; min-width: 2500px !important; }
    .custom-table th { background-color: #1f295a; color: #00f2ff; text-align: left; padding: 15px; font-size: 11px; text-transform: uppercase; position: sticky; top: 0; z-index: 99; }
    .custom-table td { padding: 12px; border-bottom: 1px solid #1f295a; font-size: 11px; color: #e0e0e0; white-space: pre-wrap !important; }
    
    .status-badge { padding: 4px 10px; border-radius: 12px; font-size: 10px; font-weight: bold; }
    .status-ativo { background-color: rgba(46, 204, 113, 0.2); color: #2ecc71; border: 1px solid #2ecc71; }
    .status-cancelado { background-color: rgba(231, 76, 60, 0.2); color: #e74c3c; border: 1px solid #e74c3c; }

    .card-hud { background: rgba(18, 22, 41, 0.7); border: 1px solid #1f295a; padding: 12px; border-radius: 10px; text-align: center; height: 100%; min-height: 110px; display: flex; flex-direction: column; justify-content: center; box-shadow: 0 4px 15px rgba(0,0,0,0.3); }
    .neon-pink { color: #ff007a; border-top: 2px solid #ff007a; }
    .neon-green { color: #2ecc71; border-top: 2px solid #2ecc71; }
    .neon-blue { color: #00f2ff; border-top: 2px solid #00f2ff; }
    .neon-purple { color: #bc13fe; border-top: 2px solid #bc13fe; }
    .neon-red { color: #ff4b4b; border-top: 2px solid #ff4b4b; }
    
    div.stButton > button { background-color: #00f2ff !important; color: #000000 !important; font-weight: bold !important; border: none !important; }
    header {visibility: hidden;} footer {visibility: hidden;}
    .logo-container { position: relative; top: -10px; left: 0px; margin-bottom: 10px; }
    </style>
    """, unsafe_allow_html=True)

# --- TELA DE LOGIN ---
if not st.session_state.logado:
    _, centro_login, _ = st.columns([1, 1.2, 1])
    with centro_login:
        st.markdown('<div class="login-box">', unsafe_allow_html=True)
        if os.path.exists(caminho_logo):
            st.image(caminho_logo, width=180)
        st.markdown("<h2 style='color: #00f2ff; margin-bottom:30px;'>ACESSO RESTRITO</h2>", unsafe_allow_html=True)
        user_in = st.text_input("USUÁRIO", key="l_u").strip()
        pass_in = st.text_input("SENHA", type="password", key="l_p").strip()
        if st.button("ENTRAR NO SISTEMA", use_container_width=True):
            df_u = carregar_usuarios()
            def limpar(v): return str(v).strip().upper().replace('.0', '')
            valido = df_u[(df_u['usuario'].apply(limpar) == user_in.upper()) & 
                          (df_u['senha'].apply(limpar) == pass_in.upper())]
            if not valido.empty:
                st.session_state.logado = True
                st.session_state.usuario_ativo = user_in
                st.session_state.nivel_ativo = str(valido.iloc[0]['nivel']).upper()
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# --- VARIÁVEL ADMIN (Criada agora para não dar erro) ---
is_admin = st.session_state.nivel_ativo == "ADMIN"

# 1. Primeiro definimos as funções (as "receitas")
def carregar_tags():
    padrao = {"tags": {}, "last_selection": {}}
    if os.path.exists("tags_salvas.json"):
        try:
            with open("tags_salvas.json", "r", encoding="utf-8") as f:
                return json.load(f)
        except: 
            return padrao
    return padrao

def salvar_tags(dados):
    try:
        with open("tags_salvas.json", "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
    except Exception as e:
        st.error(f"Erro ao salvar tags: {e}")

# 2. Depois de definidas, aí sim inicializamos o estado da sessão
if "dados_tags" not in st.session_state:
    st.session_state.dados_tags = carregar_tags()


# --- BLOCO DE FUNÇÕES MOTOR (OBRIGATÓRIO) ---
def transformar_curso(chave):
    entrada = st.session_state[chave].strip()
    if not entrada: return
    match = re.search(r'(\d+)$', entrada)
    if match:
        codigo = match.group(1); nome = DIC_CURSOS.get(codigo)
        if nome:
            base = entrada[:match.start()].strip().rstrip('+').strip()
            st.session_state[chave] = (f"{base} + {nome}" if base and nome.upper() not in base.upper() else (base if base else nome)).upper()
    else: st.session_state[chave] = entrada.upper()

def extrair_valor_geral(texto):
    if not texto: return 0.0
    try:
        v = re.findall(r'\d+(?:\.\d+)?(?:,\d+)?', str(texto).replace('.', '').replace(',', '.'))
        return float(v[0]) if v else 0.0
    except: return 0.0

def atualizar_pagamento():
    suffix = f"a_{st.session_state.reset_aluno}_{st.session_state.reset_geral}"
    base = st.session_state.get(f"f_pagto_{suffix}", "").split('|')[0].strip()
    novo = base
    if st.session_state.get(f"chk_1_{suffix}"): novo += " | Após pagamento link cartão, avisar Natália para liberação In-glês"
    if st.session_state.get(f"chk_2_{suffix}"): novo += " | Caso pague via link cartão, avisar Natália para liberação curso bônus a escolha"
    if st.session_state.get(f"chk_3_{suffix}"): novo += " | AGUARDANDO CONFIRMAÇÃO DA MATRÍCULA"
    st.session_state[f"f_pagto_{suffix}"] = novo.upper()

def reset_campos_subir():
    for c in ["in_user", "in_nome", "in_cell", "in_doc", "in_city", "in_cour", "in_pay", "in_sell", "in_date"]:
        if c in st.session_state: st.session_state[c] = ""
    st.session_state.df_final_processado = None

# --- SISTEMA PÓS-LOGIN ---
if os.path.exists(caminho_logo):
    st.markdown('<div class="logo-container">', unsafe_allow_html=True)
    st.image(caminho_logo, width=90)
    st.markdown('</div>', unsafe_allow_html=True)

# --- ESTADOS DE SESSÃO DO SISTEMA ---
if "lista_previa" not in st.session_state: st.session_state.lista_previa = []
if "reset_aluno" not in st.session_state: st.session_state.reset_aluno = 0
if "reset_geral" not in st.session_state: st.session_state.reset_geral = 0
if "df_final_processado" not in st.session_state: st.session_state.df_final_processado = None
# --- INICIALIZAÇÃO DE TAGS (COLOQUE ISSO ANTES DAS ABAS) ---
if "dados_tags" not in st.session_state:
    st.session_state.dados_tags = carregar_tags()

# --- FUNÇÕES ---
def safe_read():
    try: return conn.read(ttl="10s").dropna(how='all')
    except: return pd.DataFrame()

def formatar_cpf(chave):
    valor = re.sub(r'\D', '', st.session_state[chave])
    if len(valor) == 11: st.session_state[chave] = f"{valor[:3]}.{valor[3:6]}.{valor[6:9]}-{valor[9:]}"

DIC_CURSOS = {"00": "COLÉGIO COMBO", "1": "PREPARATÓRIO JOVEM BANCÁRIO", "2": "10 CURSOS PROFISSIONALIZANTES", "3": "PREPARATÓRIO AGRO", "4": "INGLÊS", "5": "JOVEM NO DIREITO", "6": "PRÉ MILITAR", "7": "PREPARATÓRIO ENCCEJA", "8": "JOVEM NA AVIAÇÃO", "9": "INFORMÁTICA", "10": "ADMINISTRAÇÃO"}

# --- NAVEGAÇÃO RESTRITA POR NÍVEL ---
is_admin = st.session_state.nivel_ativo == "ADMIN"
is_consulta = st.session_state.nivel_ativo == "CONSULTA"

if is_admin:
    lista_abas = ["📑 CADASTRO", "🖥️ GERENCIAMENTO", "📊 RELATÓRIOS", "📤 SUBIR ALUNOS", "👥 USUÁRIOS"]
elif is_consulta:
    lista_abas = ["🖥️ GERENCIAMENTO", "📊 RELATÓRIOS"]
else:
    lista_abas = ["🖥️ GERENCIAMENTO"]

abas = st.tabs(lista_abas)

# Definição das variáveis para evitar o erro de "não definido"
tab_cad = tab_ger = tab_rel = tab_subir = tab_users = None

if is_admin:
    tab_cad, tab_ger, tab_rel, tab_subir, tab_users = abas[0], abas[1], abas[2], abas[3], abas[4]
elif is_consulta:
    tab_ger, tab_rel = abas[0], abas[1]


# --- ABA 1: CADASTRO ---
if tab_cad: 
    with tab_cad:
        _, centro, _ = st.columns([0.2, 5.6, 0.2])
        
        with centro:
            # Suffixos para evitar conflito de IDs de widgets
            s_al = f"a_{st.session_state.reset_aluno}_{st.session_state.reset_geral}"
            s_ge = f"g_{st.session_state.reset_geral}"
            
            fields = [
                ("ID:", f"f_id_{s_al}"), 
                ("ALUNO:", f"f_nome_{s_al}"), 
                ("TEL. RESPONSÁVEL:", f"f_tel_resp_{s_al}"),
                ("TEL. ALUNO:", f"f_tel_aluno_{s_al}"), 
                ("CPF RESPONSÁVEL:", f"f_cpf_{s_al}"), 
                ("CIDADE:", f"f_cid_{s_ge}"),
                ("CURSO CONTRATADO:", f"input_curso_key_{s_al}"), 
                ("FORMA DE PAGAMENTO:", f"f_pagto_{s_al}"),
                ("VENDEDOR:", f"f_vend_{s_ge}"), 
                ("DATA DA MATRÍCULA:", f"f_data_{s_ge}")
            ]
            
            for l, k in fields:
                cl, ci = st.columns([1.2, 3.8])
                cl.markdown(f"<label>{l}</label>", unsafe_allow_html=True)
                if "curso" in k: 
                    ci.text_input(l, key=k, on_change=transformar_curso, args=(k,), label_visibility="collapsed")
                elif "f_cpf" in k: 
                    ci.text_input(l, key=k, on_change=formatar_cpf, args=(k,), label_visibility="collapsed")
                else: 
                    ci.text_input(l, key=k, label_visibility="collapsed")
            
            st.write("")
            _, c1, c2, c3, _ = st.columns([1.2, 1.2, 1.2, 1.2, 0.2])
            c1.checkbox("LIB. IN-GLÊS", key=f"chk_1_{s_al}", on_change=atualizar_pagamento)
            c2.checkbox("CURSO BÔNUS", key=f"chk_2_{s_al}", on_change=atualizar_pagamento)
            c3.checkbox("CONFIRMAÇÃO", key=f"chk_3_{s_al}", on_change=atualizar_pagamento)
            
            st.write("")
            _, b1, b2, _ = st.columns([1.2, 1.9, 1.9, 0.2])
            
            with b1:
                if st.button("💾 SALVAR ALUNO"):
                    if st.session_state[f"f_nome_{s_al}"]:
                        st.session_state.lista_previa.append({
                            "ID": st.session_state[f"f_id_{s_al}"].upper(),
                            "Aluno": st.session_state[f"f_nome_{s_al}"].upper(),
                            "Tel_Resp": str(st.session_state[f"f_tel_resp_{s_al}"]), 
                            "Tel_Aluno": str(st.session_state[f"f_tel_aluno_{s_al}"]),
                            "CPF": st.session_state[f"f_cpf_{s_al}"],
                            "Cidade": st.session_state[f"f_cid_{s_ge}"].upper(), 
                            "Course": st.session_state[f"input_curso_key_{s_al}"].upper(),
                            "Pagto": st.session_state[f"f_pagto_{s_al}"].upper(),
                            "Vendedor": st.session_state[f"f_vend_{s_ge}"].upper(),
                            "Data_Mat": st.session_state[f"f_data_{s_ge}"]
                        })
                        st.session_state.reset_aluno += 1
                        st.rerun()
                    else:
                        st.warning("Preencha pelo menos o nome do aluno.")
                        
            with b2:
                if st.button("📤 ENVIAR PLANILHA"):
                    if st.session_state.lista_previa:
                        try:
                            creds_info = st.secrets["connections"]["gsheets"]
                            client = gspread.authorize(Credentials.from_service_account_info(creds_info, scopes=["https://www.googleapis.com/auth/spreadsheets"]))
                            ws = client.open_by_url(creds_info["spreadsheet"]).get_worksheet(0)
                            
                            d_f = []
                            for a in st.session_state.lista_previa:
                                d_f.append([
                                    "ATIVO", "MGA", "A DEFINIR", 
                                    "SIM" if "10 CURSOS" in a["Course"] else "NÃO", 
                                    "A DEFINIR" if "INGLÊS" in a["Course"] else "NÃO", 
                                    date.today().strftime("%d/%m/%Y"), 
                                    a["ID"], a["Aluno"], a["Tel_Resp"], a["Tel_Aluno"], 
                                    a["CPF"], a["Cidade"], a["Course"], a["Pagto"], 
                                    a["Vendedor"], a["Data_Mat"]
                                ])
                            
                            ws.append_rows(d_f, value_input_option='RAW')
                            st.session_state.lista_previa = []
                            st.session_state.reset_geral += 1
                            st.success("Enviado com sucesso!")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao enviar: {e}")
                    else:
                        st.info("Nenhum aluno na lista de pré-visualização.")
            
            if st.session_state.lista_previa: 
                st.markdown(f"### 📋 PRÉ-VISUALIZAÇÃO ({len(st.session_state.lista_previa)} ALUNOS)")
                st.dataframe(pd.DataFrame(st.session_state.lista_previa), use_container_width=True, hide_index=True)

# --- ABA 2: GERENCIAMENTO ---
if tab_ger:
    with tab_ger:
        cf1, cf2, cf3, cf4 = st.columns([2.5, 1.5, 1.5, 0.5])
        
        with cf1: 
            bu = st.text_input("🔍 Buscar...", key="busca_ger", placeholder="Nome ou ID", label_visibility="collapsed")
        
        with cf2: 
            fs = st.selectbox("Status", ["Todos", "ATIVO", "CANCELADO"], key="filtro_status", label_visibility="collapsed")
        
        with cf3: 
            fu = st.selectbox("Unidade", ["Todos", "MGA"], key="filtro_unid", label_visibility="collapsed")
        
        with cf4: 
            if st.button("🔄", key="btn_ref"): 
                st.cache_data.clear()
                st.rerun()
                
        df_g = safe_read()
        
        if not df_g.empty:
            # Garante que as colunas existam antes de renomear para evitar erros
            try:
                df_g.columns = ['STATUS', 'UNID.', 'TURMA', '10C', 'ING', 'DT_CAD', 'ID', 'ALUNO', 'TEL_RESP', 'TEL_ALU', 'CPF', 'CIDADE', 'CURSO', 'PAGTO', 'VEND.', 'DT_MAT']
                
                # Filtros dinâmicos
                if bu: 
                    df_g = df_g[df_g['ALUNO'].str.contains(bu, case=False, na=False) | df_g['ID'].str.contains(bu, case=False, na=False)]
                if fs != "Todos": 
                    df_g = df_g[df_g['STATUS'] == fs]
                if fu != "Todos": 
                    df_g = df_g[df_g['UNID.'] == fu]
                
                rows = ""
                # Itera sobre o DataFrame invertido (mais recentes primeiro)
                for _, r in df_g.iloc[::-1].iterrows():
                    sc = "status-badge status-ativo" if r['STATUS'] == "ATIVO" else "status-badge status-cancelado"
                    rows += f"""
                    <tr>
                        <td><span class='{sc}'>{r['STATUS']}</span></td>
                        <td>{r['UNID.']}</td>
                        <td>{r['TURMA']}</td>
                        <td>{r['10C']}</td>
                        <td>{r['ING']}</td>
                        <td>{r['DT_CAD']}</td>
                        <td style='color:#00f2ff;font-weight:bold'>{r['ID']}</td>
                        <td style='color:#00f2ff;font-weight:bold'>{r['ALUNO']}</td>
                        <td>{r['TEL_RESP']}</td>
                        <td>{r['TEL_ALU']}</td>
                        <td>{r['CPF']}</td>
                        <td>{r['CIDADE']}</td>
                        <td>{r['CURSO']}</td>
                        <td>{r['PAGTO']}</td>
                        <td>{r['VEND.']}</td>
                        <td>{r['DT_MAT']}</td>
                    </tr>"""
                
                # Montagem da tabela em HTML/CSS
                st.markdown(f'''
                    <div class="custom-table-wrapper">
                        <table class="custom-table">
                            <thead>
                                <tr>{''.join([f'<th>{h}</th>' for h in df_g.columns])}</tr>
                            </thead>
                            <tbody>
                                {rows}
                            </tbody>
                        </table>
                    </div>
                ''', unsafe_allow_html=True)
            except Exception as e:
                st.error(f"Erro ao processar dados de gerenciamento: {e}")

# --- ABA 3: RELATÓRIOS ---
if tab_rel:
    with tab_rel:
        df_r = safe_read()
        if not df_r.empty:
            # Limpeza de nomes de colunas
            df_r.columns = [c.strip() for c in df_r.columns]
            
            dt_col = "Data Matrícula"
            df_r[dt_col] = pd.to_datetime(df_r[dt_col], dayfirst=True, errors='coerce')
            
            # Filtro de data
            iv = st.date_input("Filtrar Período (Data de Matrícula)", 
                              value=(date.today()-timedelta(days=7), date.today()), 
                              format="DD/MM/YYYY")
            
            if len(iv) == 2:
                df_f = df_r.loc[(df_r[dt_col].dt.date >= iv[0]) & (df_r[dt_col].dt.date <= iv[1])].copy()
                
                # --- LÓGICA DE PROCESSAMENTO (RESILIENTE) ---
                v_taxa = 0.0
                v_cartao = 0.0
                v_entrada = 0.0
                pagamentos = df_f['Pagamento'].tolist()

                for linha in pagamentos:
                    if not linha or str(linha).strip() == "": continue
                    linha_upper = str(linha).upper()
                    
                    # 1. TRATAMENTO DE ALTERAÇÕES
                    if "ALTERAÇÃO" in linha_upper or "ALTEROU PARA" in linha_upper:
                        match_alt = re.search(r'\((?:.*?PARA\s+)?(.*?)\)', linha_upper)
                        if match_alt:
                            linha_upper = match_alt.group(1)

                    # 2. CAPTURA DE TAXAS
                    taxas_na_linha = re.findall(r'TAXA.*?(\d+)', linha_upper)
                    for t in taxas_na_linha:
                        try: v_taxa += float(t)
                        except: pass
                    if "TAXA" in linha_upper and "PAGA" in linha_upper and not taxas_na_linha:
                        v_taxa += 50.0

                    # 3. REGRA DO CARTÃO
                    match_mult = re.findall(r'(\d+)\s*[X]\s*(?:R\$)?\s*([\d\.,]+)', linha_upper)
                    if match_mult and ("CARTÃO" in linha_upper or "LINK" in linha_upper):
                        for qtd, val in match_mult:
                            try:
                                v_u = val.replace('.', '').replace(',', '.')
                                v_cartao += int(qtd) * float(v_u)
                            except: pass
                    else:
                        match_fixo = re.findall(r'(?:PAGO|R\$)\s*([\d\.]+,\d{2}|[\d\.]+)', linha_upper)
                        for val in match_fixo:
                            try:
                                v_l = val.replace('.', '').replace(',', '.')
                                valor_limpo = float(v_l)
                                if valor_limpo != 50.0:
                                    if "CARTÃO" in linha_upper or "LINK" in linha_upper:
                                        v_cartao += valor_limpo
                                    else:
                                        v_entrada += valor_limpo
                            except: pass

                    # 4. ENTRADAS DIVERSAS
                    if any(x in linha_upper for x in ["BOLETO", "PIX", "DINHEIRO", "DÉBITO"]):
                        match_ent = re.findall(r'(?:PARCELA|ENTRADA|PIX|DINHEIRO|DÉBITO).*?(?:R\$)?\s*([\d\.,]+)', linha_upper)
                        for val in match_ent:
                            try:
                                v_e = val.replace('.', '').replace(',', '.')
                                valor_ent = float(v_e)
                                if valor_ent not in [float(v.replace('.', '').replace(',', '.')) for v in re.findall(r'(?:PAGO|R\$)\s*([\d\.]+,\d{2}|[\d\.]+)', linha_upper) if v]:
                                    v_entrada += valor_ent
                            except: pass
                
                total_final = v_taxa + v_cartao + v_entrada

                # --- DASHBOARD (CARDS) ---
                c1, c2, c3, c4, c5, c6 = st.columns(6)
                with c1: st.markdown(f'<div class="card-hud neon-pink"><span class="stat-label">MATRÍCULAS</span><h2>{len(df_f)}</h2></div>', unsafe_allow_html=True)
                with c2: st.markdown(f'<div class="card-hud neon-green"><span class="stat-label">ATIVOS</span><h2>{len(df_f[df_f["STATUS"].str.upper()=="ATIVO"])}</h2></div>', unsafe_allow_html=True)
                with c3: st.markdown(f'<div class="card-hud neon-red"><span class="stat-label">CANCELADOS</span><h2>{len(df_f[df_f["STATUS"].str.upper()=="CANCELADO"])}</h2></div>', unsafe_allow_html=True)
                with c4: st.markdown(f'<div class="card-hud neon-blue"><span class="stat-label">TOTAL RECEBIDO</span><h2 style="font-size:22px">R${total_final:,.2f}</h2></div>', unsafe_allow_html=True)
                
                with c5:
                    df_f['v_tic'] = df_f['Pagamento'].apply(extrair_valor_geral)
                    tm_b = df_f[df_f['Pagamento'].str.contains('BOLETO', na=False, case=False)]['v_tic'].mean() or 0.0
                    tm_c = df_f[df_f['Pagamento'].str.contains('CARTÃO|LINK', na=False, case=False)]['v_tic'].mean() or 0.0
                    st.markdown(f'<div class="card-hud neon-purple"><span class="stat-label">TICKET MÉDIO</span><div style="font-size:18px; font-weight:bold; color:#e0e0e0;">BOL: R${tm_b:.0f}<br>CAR: R${tm_c:.0f}</div></div>', unsafe_allow_html=True)
                
                with c6:
                    c_banc = len(df_f[df_f["Curso"].str.contains("BANCÁRIO", case=False, na=False)])
                    c_agro = len(df_f[df_f["Curso"].str.contains("AGRO", case=False, na=False)])
                    c_ing = len(df_f[df_f["Curso"].str.contains("INGLÊS", case=False, na=False)])
                    c_tec = len(df_f[df_f["Curso"].str.contains("TECNOLOGIA|INFORMÁTICA", case=False, na=False)])
                    st.markdown(f'''<div class="card-hud neon-blue"><span class="stat-label">POR ÁREA</span><div style="font-size:15px; text-align:left; color:#e0e0e0; line-height:1.4; padding-left:5px;">BANC: <b style="color:#00f2ff;">{c_banc}</b> | AGRO: <b style="color:#00f2ff;">{c_agro}</b><br>INGL: <b style="color:#00f2ff;">{c_ing}</b> | TECN: <b style="color:#00f2ff;">{c_tec}</b></div></div>''', unsafe_allow_html=True)

                st.write("")
                if len(df_f) > 0:
                    at_c = len(df_f[df_f["STATUS"].str.upper()=="ATIVO"])
                    can_c = len(df_f[df_f["STATUS"].str.upper()=="CANCELADO"])
                    fig_status = go.Figure()
                    fig_status.add_trace(go.Bar(y=["STATUS"], x=[at_c], orientation='h', marker=dict(color='#2ecc71'), text=[f"<b>ATIVOS: {at_c}</b>"], textposition='inside'))
                    fig_status.add_trace(go.Bar(y=["STATUS"], x=[can_c], orientation='h', marker=dict(color='#ff4b4b'), text=[f"<b>CANCELADOS: {can_c}</b>"], textposition='inside'))
                    fig_status.update_layout(barmode='stack', showlegend=False, height=40, margin=dict(t=5, b=5, l=10, r=10), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', xaxis=dict(showgrid=False, showticklabels=False), yaxis=dict(showgrid=False, showticklabels=False))
                    st.plotly_chart(fig_status, use_container_width=True, config={'displayModeBar': False})

                st.write("---")
                col_graf_1, col_graf_2 = st.columns(2)
                with col_graf_1:
                    st.markdown("<h4 style='text-align:center; color:#00f2ff;'>📍 CIDADES E VENDEDORES</h4>", unsafe_allow_html=True)
                    df_city_full = df_f.copy()
                    df_city_full["Vendedor_Limpo"] = df_city_full["Vendedor"].str.split(" - ").str[0].str.strip()
                    top_cities = df_city_full['Cidade'].value_counts().head(5).index
                    df_city_vends = []
                    for city in top_cities:
                        vends = df_city_full[df_city_full['Cidade'] == city]['Vendedor_Limpo'].unique()
                        vends_str = ", ".join(list(vends))
                        count = len(df_city_full[df_city_full['Cidade'] == city])
                        df_city_vends.append({"Cidade": city, "Qtd": count, "Vendedores": vends_str})
                    df_city_plot = pd.DataFrame(df_city_vends)
                    fig_city = go.Figure(go.Bar(x=df_city_plot['Cidade'], y=df_city_plot['Qtd'], text=df_city_plot.apply(lambda r: f"<b>{r['Qtd']}</b><br><span style='font-size:11px; color:#ff007a;'>{r['Vendedores']}</span>", axis=1), textposition='outside', marker=dict(color=df_city_plot['Qtd'], colorscale=[[0, '#1f295a'], [1, '#00f2ff']])))
                    fig_city.update_layout(template="plotly_dark", paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', height=450, xaxis=dict(showgrid=False), yaxis=dict(showgrid=False, showticklabels=False))
                    st.plotly_chart(fig_city, use_container_width=True)

                with col_graf_2:
                    st.markdown("<h4 style='text-align:center; color:#bc13fe;'>⚡ PERFORMANCE DE VENDAS</h4>", unsafe_allow_html=True)
                    df_temp = df_f.copy()
                    df_temp["Vendedor"] = df_temp["Vendedor"].str.split(" - ").str[0].str.strip()
                    df_stats = df_temp["Vendedor"].value_counts().reset_index().head(5)
                    df_stats.columns = ['Vendedor', 'Total']
                    fig_vend = go.Figure(go.Scatter(x=df_stats['Vendedor'], y=df_stats['Total'], mode='lines+markers+text', text=df_stats['Total'], textposition="top center", line=dict(color='#bc13fe', width=4), fill='tozeroy', fillcolor='rgba(188, 19, 254, 0.2)'))
                    fig_vend.update_layout(template="plotly_dark", paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', height=400, xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="rgba(255,255,255,0.05)", showticklabels=False))
                    st.plotly_chart(fig_vend, use_container_width=True)


# --- ABA 4: SUBIR ALUNOS ---
if tab_subir:
    with tab_subir:
        st.markdown("### 📤 IMPORTAÇÃO EAD")
        
        modo = st.radio("Método:", ["MANUAL", "AUTOMÁTICO"], horizontal=True)
        st.write("---")
        
        if modo == "AUTOMÁTICO":
            df_m = safe_read()
            if not df_m.empty:
                try:
                    col_f = df_m.columns[5]
                    df_m[col_f] = pd.to_datetime(df_m[col_f], dayfirst=True, errors='coerce')
                    data_sel = st.date_input("Filtrar Cadastro (Coluna F):", value=date.today())
                    df_filtrado = df_m[df_m[col_f].dt.date == data_sel]
                    
                    if not df_filtrado.empty:
                        cids = sorted(df_filtrado[df_m.columns[11]].unique())
                        sel_cids = st.multiselect("Cidades:", cids)
                        st.session_state.df_auto_ready = df_filtrado[df_filtrado[df_m.columns[11]].isin(sel_cids)]
                        st.info(f"{len(st.session_state.df_auto_ready)} alunos encontrados.")
                except: 
                    st.error("Erro ao processar colunas da planilha automática.")
        else:
            c1, c2 = st.columns(2)
            with c1:
                u_user = st.text_area("IDs", height=100, key="in_user")
                u_cell = st.text_area("Celulares", height=100, key="in_cell")
                u_city = st.text_area("Cidades", height=100, key="in_city")
                u_pay = st.text_area("Pagamentos", height=100, key="in_pay")
            with c2:
                u_nome = st.text_area("Nomes", height=100, key="in_nome")
                u_doc = st.text_area("Documentos", height=100, key="in_doc")
                u_cour = st.text_area("Cursos", height=100, key="in_cour")
                u_sell = st.text_area("Vendedores", height=100, key="in_sell")
            u_date = st.text_area("Datas", height=100, key="in_date")

        with st.expander("🛠️ CONFIGURAR TAGS", expanded=False):
            cursos_tags = ['PREPARATÓRIO JOVEM BANCÁRIO', 'PREPARATÓRIO AGRO', 'JOVEM NO DIREITO', 'INGLÊS', 'PRÉ MILITAR', 'ADMINISTRATIVO', 'INFORMÁTICA', 'PREPARATÓRIO ENCCEJA', 'JOVEM NA AVIAÇÃO', 'TECNOLOGIA']
            cols = st.columns(3)
            selected_tags = {}
            for i, curso in enumerate(cursos_tags):
                with cols[i % 3]:
                    st.markdown(f"<p style='font-size:10px; margin-bottom:2px; color:#00f2ff; font-weight:bold;'>{curso}</p>", unsafe_allow_html=True)
                    tags_lista = st.session_state.dados_tags.get("tags", {}).get(curso, [])
                    last_sel = st.session_state.dados_tags.get("last_selection", {}).get(curso, "")
                    idx_default = (tags_lista.index(last_sel) + 1) if last_sel in tags_lista else 0
                    c_sel, c_del = st.columns([0.4, 0.6])
                    cur_tag = c_sel.selectbox("", [""] + tags_lista, index=idx_default, key=f"sel_{curso}", label_visibility="collapsed")
                    
                    if cur_tag != last_sel:
                        st.session_state.dados_tags["last_selection"][curso] = cur_tag
                        salvar_tags(st.session_state.dados_tags)
                    
                    if c_del.button("🗑️", key=f"del_{curso}"):
                        if cur_tag and cur_tag in st.session_state.dados_tags["tags"][curso]:
                            st.session_state.dados_tags["tags"][curso].remove(cur_tag)
                            st.session_state.dados_tags["last_selection"][curso] = ""
                            salvar_tags(st.session_state.dados_tags)
                            st.rerun()
                    
                    c_new, _ = st.columns([0.4, 0.6])
                    new_tag = c_new.text_input("", placeholder="Nova...", key=f"new_{i}", label_visibility="collapsed").upper()
                    if new_tag and new_tag not in tags_lista:
                        if "tags" not in st.session_state.dados_tags: st.session_state.dados_tags["tags"] = {}
                        if curso not in st.session_state.dados_tags["tags"]: st.session_state.dados_tags["tags"][curso] = []
                        st.session_state.dados_tags["tags"][curso].append(new_tag)
                        st.session_state.dados_tags["last_selection"][curso] = new_tag
                        salvar_tags(st.session_state.dados_tags)
                        st.rerun()
                    final_tag = (new_tag if new_tag else cur_tag).upper()
                    selected_tags[curso] = final_tag

        if st.button("🚀 PROCESSAR DADOS", use_container_width=True):
            raw_list = []
            if modo == "MANUAL":
                l_ids = u_user.strip().split('\n')
                l_nomes = u_nome.strip().split('\n')
                l_pays = u_pay.strip().split('\n')
                l_cours = u_cour.strip().split('\n')
                l_cells = u_cell.strip().split('\n')
                l_docs = u_doc.strip().split('\n')
                l_citys = u_city.strip().split('\n')
                l_sells = u_sell.strip().split('\n')
                l_dates = u_date.strip().split('\n')
                min_len = len(l_ids)
                if min_len > 0:
                    for i in range(min_len):
                        try:
                            raw_list.append({
                                "User": l_ids[i], 
                                "Nome": l_nomes[i] if i < len(l_nomes) else "", 
                                "Pay": l_pays[i] if i < len(l_pays) else "", 
                                "Cour": l_cours[i] if i < len(l_cours) else "", 
                                "Cell": l_cells[i] if i < len(l_cells) else "", 
                                "Doc": l_docs[i] if i < len(l_docs) else "", 
                                "City": l_citys[i] if i < len(l_citys) else "", 
                                "Sell": l_sells[i] if i < len(l_sells) else "", 
                                "Date": l_dates[i] if i < len(l_dates) else ""
                            })
                        except: continue
            elif "df_auto_ready" in st.session_state and st.session_state.df_auto_ready is not None:
                for _, r in st.session_state.df_auto_ready.iterrows():
                    raw_list.append({"User": r.iloc[6], "Nome": r.iloc[7], "Cell": r.iloc[9], "Doc": r.iloc[10], "City": r.iloc[11], "Cour": r.iloc[12], "Pay": r.iloc[13], "Sell": r.iloc[14], "Date": r.iloc[15]})
            
            if raw_list:
                try:
                    wb_c = load_workbook(ARQUIVO_CIDADES)
                    ws_c = wb_c.active
                    c_map = {str(r[1]).strip().upper(): str(r[2]) for r in ws_c.iter_rows(min_row=2, values_only=True) if r[1]}
                except: c_map = {}
                
                processed = []
                for item in raw_list:
                    c_orig = str(item['Cour']).upper()
                    p_orig = str(item['Pay']).upper()
 # 1. Encontrar a posição de cada curso dentro da string original
posicoes_tags = []
for curso_nome, tag_valor in selected_tags.items():
    if tag_valor: # Só processa se houver uma tag definida
        pos = c_orig.find(curso_nome)
        if pos != -1:
            posicoes_tags.append((pos, tag_valor))

# 2. Ordenar as tags pela posição que o curso aparece no texto original
posicoes_tags.sort() # Ordena pelo primeiro item da tupla (a posição)

# 3. Extrair apenas as tags já ordenadas e juntar com ponto
tags_f = [t[1] for t in posicoes_tags]
c_final = ".".join(tags_f).upper() if tags_f else c_orig

                    
                    p_final = "PENDENTE"
                    has_bol = "BOLETO" in p_orig
                    has_car = "CARTÃO" in p_orig or "LINK" in p_orig
                    if (has_bol and not has_car): p_final = "BOLETO"
                    elif (has_car and not has_bol): p_final = "CARTÃO"
                    
                    obs_final = f"{c_final} | {c_orig} | {p_orig}".upper()
                    ouro_val = "1" if "10 CURSOS PROFISSIONALIZANTES" in obs_final else "0"
                    
                    processed.append({
                        "username": item['User'], "email2": f"{item['User']}@profissionalizaead.com.br", 
                        "name": str(item['Nome']).split(" ")[0].upper(), 
                        "lastname": " ".join(str(item['Nome']).split(" ")[1:]).upper(), 
                        "cellphone2": str(item['Cell']), "document": item['Doc'], 
                        "city2": c_map.get(str(item['City']).upper(), item['City']), 
                        "courses": c_final, "payment": p_final, "observation": obs_final, 
                        "ouro": ouro_val, "password": "futuro", "role": "1", 
                        "secretary": "MGA", "seller": item['Sell'], "contract_date": item['Date'], "active": "1"
                    })
                st.session_state.df_final_processado = pd.DataFrame(processed)

        if st.session_state.df_final_processado is not None:
            df = st.session_state.df_final_processado
            mask = df['payment'] == "PENDENTE"
            if mask.any():
                st.warning("⚠️ Confirmação necessária:")
                df_conf = df.loc[mask, ["username", "name", "observation"]].copy()
                df_conf.columns = ["ID", "Nome", "Texto Original (Pagamento)"]
                df_conf["Forma Final"] = "BOLETO"
                edited = st.data_editor(df_conf, column_config={"Forma Final": st.column_config.SelectboxColumn("Forma", options=["BOLETO", "CARTÃO"], required=True)}, disabled=["ID", "Nome", "Texto Original (Pagamento)"], hide_index=True, use_container_width=True, key="pag_editor")
                if st.button("✅ CONFIRMAR E GERAR EXCEL"):
                    for _, row in edited.iterrows(): 
                        df.loc[df['username'] == row["ID"], "payment"] = row["Forma Final"]
                    st.session_state.df_final_processado = df
                    st.rerun()
            
            if not (st.session_state.df_final_processado['payment'] == "PENDENTE").any():
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.append(list(st.session_state.df_final_processado.columns))
                for r in st.session_state.df_final_processado.values.tolist(): 
                    ws.append([str(val) for val in r])
                wb.save(output)
                st.download_button("📥 BAIXAR EXCEL FINAL", output.getvalue(), f"ead_{date.today()}.xlsx", on_click=reset_campos_subir, use_container_width=True)

# --- ABA USUÁRIOS (Corrigida com Scopes) ---
if tab_users: 
    with tab_users:
        st.markdown("### 👥 GESTÃO DE ACESSOS")
        
        with st.form("novo_user_final", clear_on_submit=True):
            col1, col2, col3 = st.columns(3)
            nu = col1.text_input("Novo Usuário").strip()
            ns = col2.text_input("Senha").strip()
            nv = col3.selectbox("Nível", ["ADMIN", "CONSULTA"])
            
            if st.form_submit_button("CADASTRAR"):
                if nu and ns:
                    try:
                        c_info = st.secrets["connections"]["gsheets"]
                        sc = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
                        creds = Credentials.from_service_account_info(c_info, scopes=sc)
                        client = gspread.authorize(creds)
                        ws = client.open_by_url(c_info["spreadsheet"]).worksheet("usuários")
                        ws.append_row([nu, ns, nv])
                        st.success("Cadastrado!")
                        st.cache_data.clear()
                        st.rerun()
                    except Exception as e: 
                        st.error(f"Erro: {e}")

        st.write("---")
        # Mostra a tabela de usuários cadastrados
        st.dataframe(carregar_usuarios(), use_container_width=True, hide_index=True)

# --- SIDEBAR (BARRA LATERAL) ---
# A Sidebar fica fora dos 'ifs' das abas para aparecer em todas as telas
with st.sidebar:
    st.write(f"Logado: **{st.session_state.usuario_ativo}**")
    st.write(f"Nível: **{st.session_state.nivel_ativo}**")
    if st.button("SAIR"):
        st.session_state.logado = False
        st.rerun()
